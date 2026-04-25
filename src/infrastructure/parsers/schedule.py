from typing import Optional

import openpyxl

from src.core.dto import DayDto, LessonDto, ScheduleDto
from src.core.enums import ScheduleType, WeekDay


class ExcelParser:
    def __init__(self, path):
        self.file = openpyxl.load_workbook(path, data_only=True)
        self.sheets = self.file.worksheets

    def parse_all(self) -> list[ScheduleDto]:
        return (self.students()) + (self.teachers()) + (self.rooms())

    def students(
        self, count: Optional[int] = None, min_row: int = 1, min_column: int = 1
    ) -> list[ScheduleDto]:
        sheet = self.sheets[0]
        schedule_list: list[ScheduleDto] = []
        if count is None:
            count = sheet.max_column // 2
        for grade_number in range(count):
            schedule = ScheduleDto(
                type=ScheduleType.COMMON,
                grade=sheet.cell(min_row, min_column + (2 * grade_number)).value,
            )
            for day_number in range(5):
                day = DayDto(name=WeekDay(day_number + 1))
                for lesson_number in range(8):
                    row = (min_row + 1) + 2 * lesson_number + day_number * 8 * 2
                    column = min_column + 2 * grade_number
                    lesson1 = sheet.cell(row, column).value
                    room1 = sheet.cell(row, column + 1).value
                    lesson2 = sheet.cell(row + 1, column).value
                    room2 = sheet.cell(row + 1, column + 1).value
                    if lesson2 is None and room2:
                        if room1:
                            day.lessons.append(
                                LessonDto(number=(lesson_number + 1), name=lesson1, room=str(room1))
                            )
                        day.lessons.append(
                            LessonDto(number=(lesson_number + 1), name=lesson1, room=str(room2))
                        )
                    elif lesson2:
                        day.lessons.append(
                            LessonDto(number=(lesson_number + 1), name=lesson1, room=str(room1))
                        )
                        day.lessons.append(
                            LessonDto(number=(lesson_number + 1), name=lesson2, room=str(room2))
                        )
                    else:
                        day.lessons.append(
                            LessonDto(number=(lesson_number + 1), name=None, room=None)
                        )
                schedule.days.append(day)
            schedule_list.append(schedule)
        return schedule_list

    def teachers(
        self, count: Optional[int] = None, min_row: int = 1, min_column: int = 1
    ) -> list[ScheduleDto]:
        sheet = self.sheets[1]
        schedule_list: list[ScheduleDto] = []
        if count is None:
            count = sheet.max_row // 2
        for teacher in range(count):
            schedule = ScheduleDto(
                type=ScheduleType.TEACHER,
                grade=sheet.cell(min_row + (2 * teacher), min_column).value,
            )
            for day_number in range(5):
                day = DayDto(name=WeekDay(day_number + 1))
                for lesson_number in range(1, 8 + 1):
                    row = min_row + 2 * teacher
                    column = (min_column + 1) + 2 * (lesson_number - 1) + day_number * 8 * 2
                    grade = sheet.cell(row, column).value
                    group = sheet.cell(row, column + 1).value
                    room2 = sheet.cell(row + 1, column + 1).value
                    if grade and group is None:
                        day.lessons.append(
                            LessonDto(number=lesson_number, name=str(grade), room=str(room2))
                        )
                    elif grade:
                        day.lessons.append(
                            LessonDto(
                                number=lesson_number, name=f"{grade}[{group}]", room=str(room2)
                            )
                        )
                    else:
                        day.lessons.append(LessonDto(number=lesson_number, name=None, room=None))
                schedule.days.append(day)
            schedule_list.append(schedule)
        return schedule_list

    def rooms(
        self, count: Optional[int] = None, min_row: int = 1, min_column: int = 1
    ) -> list[ScheduleDto]:
        sheet = self.sheets[2]
        schedule_list: list[ScheduleDto] = []
        if count is None:
            count = sheet.max_row
        for room in range(count):
            schedule = ScheduleDto(
                type=ScheduleType.ROOM, grade=str(sheet.cell(min_row + room, min_column).value)
            )
            for day_number in range(6):
                day = DayDto(name=WeekDay(day_number + 1))
                for lesson_number in range(1, 10 + 1):
                    row = min_row + room
                    column = (min_column + 1) + (lesson_number - 1) + day_number * 10
                    grade = sheet.cell(row, column).value
                    name = sheet.cell(min_row + room, min_column).value
                    day.lessons.append(
                        LessonDto(
                            number=lesson_number,
                            name=str(grade) if grade else None,
                            room=str(name) if name else None,
                        )
                    )
                schedule.days.append(day)
            schedule_list.append(schedule)
        return schedule_list
