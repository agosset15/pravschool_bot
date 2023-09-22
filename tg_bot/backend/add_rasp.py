from db.methods.create import create_schedule
import openpyxl


class Exel:
    def __init__(self, xl_file, dira):
        wb = f'{dira}/tg_bot/xl_uploads/{xl_file}'
        self.file_to_read = openpyxl.load_workbook(wb, data_only=True)

    def rasp(self, min_row: int, max_row: int):
        sheet = self.file_to_read['классы']
        data = []
        for i in range(4, 33, 4):
            les1 = []
            clas1 = []
            les2 = []
            clas2 = []
            for col1 in range(min_row, max_row + 1, 2):
                value = sheet.cell(col1, i).value
                les1.append(value)
                value = sheet.cell(col1, i + 1).value
                clas1.append(value)
            for col2 in range(min_row + 1, max_row + 1, 2):
                value = sheet.cell(col2, i).value
                les2.append(value)
                value = sheet.cell(col2, i + 1).value
                clas2.append(value)
            for q in range(len(les1)):
                if les1[q] is None:
                    text = " "
                    data.append(text)
                else:
                    if les2[q] is None and clas1[q]:
                        text = f"{les1[q]}({clas1[q]}/{clas2[q]})"
                        data.append(text)
                    elif les2[q] is None and clas1[q] is None:
                        text = f"{les1[q]}({clas2[q]})"
                        data.append(text)
                    else:
                        text = f"{les1[q]}({clas1[q]})//{les2[q]}({clas2[q]})"
                        data.append(text)
        main_data = []
        for w in range(0, len(data), 8):
            main_data.append([f"1.{data[w]}", f"2.{data[w + 1]}", f"3.{data[w + 2]}", f"4.{data[w + 3]}",
                              f"5.{data[w + 4]}", f"6.{data[w + 5]}", f"7.{data[w + 6]}", f"8.{data[w + 7]}"])

        ids = []
        for i in range(1, 10):
            for day in range(1, 6):
                ids.append(f"{i}_{day}")
        for i in range(1, 7):
            for day in range(1, 6):
                if i == 1:
                    ids.append(f'101_{day}')
                if i == 2:
                    ids.append(f'102_{day}')
                if i == 3:
                    ids.append(f'103_{day}')
                if i == 4:
                    ids.append(f"111_{day}")
                if i == 5:
                    ids.append(f"112_{day}")
                if i == 6:
                    ids.append(f'113_{day}')

        for s in range(0, len(main_data)):
            day = ids[s].split('_')[1]
            clas = ids[s].split('_')[0]
            create_schedule(int(clas), int(day), str(main_data[s]), False, False)

    def tchr_rasp(self, tchr_count: int, min_row: int, max_row: int):
        sheet = self.file_to_read['учителя']
        data = []
        for i in range(min_row, max_row + 1, 2):
            data1 = []
            data2 = []
            for col in range(3, 83, 2):
                value = sheet.cell(i, col).value
                data1.append(value)
            for col1 in range(4, 83, 2):
                value1 = sheet.cell(i + 1, col1).value
                data2.append(value1)
            for q in range(len(data1)):
                if data1[q] is None:
                    text = " "
                else:
                    text = f"{data1[q]}({data2[q]})"
                data.append(text)
        main_data = []
        for w in range(0, len(data), 8):
            main_data.append([f"1.{data[w]}", f"2.{data[w + 1]}", f"3.{data[w + 2]}", f"4.{data[w + 3]}",
                              f"5.{data[w + 4]}", f"6.{data[w + 5]}", f"7.{data[w + 6]}", f"8.{data[w + 7]}"])

        ids = []
        for i in range(1, tchr_count + 1):
            for day in range(1, 6):
                ids.append(f"{i}_{day}")

        for s in range(0, len(main_data)):
            clas = ids[s].split('_')[0]
            day = ids[s].split('_')[1]
            create_schedule(int(clas), int(day), str(main_data[s]), True, False)

    def kab(self, kab_count: int, min_row: int, max_row: int):
        sheet = self.file_to_read['кабинеты']
        data = []
        for i in range(min_row, max_row + 1):
            data1 = []
            for col in range(3, 43):
                value = sheet.cell(i, col).value
                data1.append(value)
            for q in range(len(data1)):
                if data1[q] is None:
                    text = " "
                else:
                    text = f"{data1[q]}"
                data.append(text)
        main_data = []
        for w in range(0, len(data), 8):
            main_data.append([f"1.{data[w]}", f"2.{data[w + 1]}", f"3.{data[w + 2]}", f"4.{data[w + 3]}",
                              f"5.{data[w + 4]}", f"6.{data[w + 5]}", f"7.{data[w + 6]}", f"8.{data[w + 7]}"])

        ids = []
        for i in range(1, kab_count + 1):
            for day in range(1, 6):
                ids.append(f'{i}_{day}')

        for s in range(0, len(main_data)):
            clas = ids[s].split('_')[0]
            day = ids[s].split('_')[1]
            create_schedule(int(clas), int(day), str(main_data[s]), True, True)
