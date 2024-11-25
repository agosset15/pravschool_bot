import openpyxl

from tg_bot.models import DefaultService, Schedule, Day, Lesson
from tg_bot.config import days


class ExcelParser:
    def __init__(self, path, db: DefaultService):
        self.file = openpyxl.load_workbook(path, data_only=True)
        self.sheets = self.file.worksheets
        self.db = db

    async def students(self, count: int, min_row: int, min_column: int):
        sheet = self.sheets[0]
        for grade in range(count):
            schedule_id = await self.db.create(Schedule, entity=0,
                                               grade=sheet.cell(min_row-1, min_column+(2*grade)).value)
            for day in range(5):
                day_id = await self.db.create(Day, schedule_id=schedule_id, name=days[day])
                for lesson in range(1, 8+1):
                    row = min_row + 2*(lesson-1) + day * 8
                    column = min_column + 2 * grade
                    lesson1 = sheet.cell(row, column).value
                    room1 = sheet.cell(row, column+1).value
                    lesson2 = sheet.cell(row+1, column).value
                    room2 = sheet.cell(row+1, column+1).value
                    if lesson2 is None and room2:
                        if room1:
                            await self.db.create(Lesson, day_id=day_id, number=lesson, name=lesson1, room=str(room1))
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=lesson1, room=str(room2))
                    elif lesson2:
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=lesson1, room=str(room1))
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=lesson2, room=str(room2))
                    else:
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=None, room=None)

    async def teachers(self, count: int, min_row: int, min_column: int):
        sheet = self.sheets[1]
        for teacher in range(count):
            schedule_id = await self.db.create(Schedule, entity=1,
                                               grade=sheet.cell(min_row+(2*teacher), min_column-1).value)
            for day in range(5):
                day_id = await self.db.create(Day, schedule_id=schedule_id, name=days[day])
                for lesson in range(1, 8+1):
                    row = min_row + 2 * teacher
                    column = min_column + 2*(lesson-1) + day * 8
                    grade = sheet.cell(row, column).value
                    group = sheet.cell(row, column + 1).value
                    room2 = sheet.cell(row+1, column+1).value
                    if grade and group is None:
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=grade, room=str(room2))
                    elif grade:
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=f"{grade}[{group}]", room=str(room2))
                    else:
                        await self.db.create(Lesson, day_id=day_id, number=lesson, name=None, room=None)

    async def rooms(self, count: int, min_row: int, min_column: int):
        sheet = self.sheets[2]
        for room in range(count):
            schedule_id = await self.db.create(Schedule, entity=2,
                                               grade=str(sheet.cell(min_row+room, min_column-1).value))
            for day in range(6):
                day_id = await self.db.create(Day, schedule_id=schedule_id, name=(days+['Суббота'])[day])
                for lesson in range(1, 10 + 1):
                    row = min_row + room
                    column = min_column + (lesson - 1) + day * 8
                    grade = sheet.cell(row, column).value
                    await self.db.create(Lesson, day_id=day_id, number=lesson, name=grade,
                                         room=str(sheet.cell(min_row+room, min_column-1).value))
