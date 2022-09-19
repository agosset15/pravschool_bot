import sqlite3
import openpyxl

class Exel:
    def __init__(self, xl_file, dir):
        base_name = 'rs-bot.db'
        wb = f'{dir}\\xl_uploads\\{xl_file}'
        self.connect = sqlite3.connect(dir + '/' + base_name)
        self.cursor = self.connect.cursor()

        self.file_to_read = openpyxl.load_workbook(wb, data_only=True)

        self.connect.commit()

    def rasp(self, min_row: int, max_row: int):
        sheet = self.file_to_read['классы']
        data = []
        for i in range(4, 61, 4):
            data1 = []
            data2 = []
            data3 = []
            data4 = []
            for col1 in range(min_row, max_row+1, 2):
                value = sheet.cell(col1, i).value
                data1.append(value)
            for col2 in range(min_row+1, max_row+1, 2):
                value = sheet.cell(col2, i + 1).value
                data2.append(value)
            for col3 in range(min_row, max_row+1, 2):
                value = sheet.cell(col3, i + 2).value
                data3.append(value)
            for col4 in range(min_row+1, max_row+1, 2):
                value = sheet.cell(col4, i + 3).value
                data4.append(value)
            print(data1)
            print(data2)
            for q in range(len(data1)):
                if data1[q] is None:
                    text = " "
                    data.append(text)
                else:
                    if data3[q] == '/':
                        text = f"{data1[q]}({data2[q]}/{data4[q]})"
                        data.append(text)
                    elif data3[q] is None:
                        text = f"{data1[q]}({data2[q]})"
                        data.append(text)
                    elif data3[q].startswith('/'):
                        m = data3[q].replace('/', '')
                        text = f"{data1[q]}({data2[q]})//{m}({data4[q]})"
                        data.append(text)
            print(data)
        main_data = []
        for w in range(0, len(data), 8):
            txt = f"1.{data[w]}" \
                  f"\n2.{data[w + 1]}" \
                  f"\n3.{data[w + 2]}" \
                  f"\n4.{data[w + 3]}" \
                  f"\n5.{data[w + 4]}" \
                  f"\n6.{data[w + 5]}" \
                  f"\n7.{data[w + 6]}" \
                  f"\n8.{data[w + 7]}"
            main_data.append(txt)

        ids = []
        for i in range(1, 10):
            for day in range(1, 6):
                a = i + day / 10
                ids.append(a)
        for i in range(1, 7):
            for day in range(1, 6):
                if i == 1:
                    ids.append(101 + day / 10)
                if i == 2:
                    ids.append(102 + day / 10)
                if i == 3:
                    ids.append(103 + day / 10)
                if i == 4:
                    ids.append(111 + day / 10)
                if i == 5:
                    ids.append(112 + day / 10)
                if i == 6:
                    ids.append(113 + day / 10)

        for s in range(0, len(main_data)):
            self.cursor.execute(f"INSERT INTO rasp VALUES (:id_day, :rasp);",
                                {'id_day': ids[s], 'rasp': main_data[s]})

        self.connect.commit()

    def tchr_rasp(self, tchr_count: int, min_row: int, max_row: int):
        sheet = self.file_to_read['учителя']
        data = []
        for i in range(min_row, max_row + 1, 2):
            data1 = []
            data2 = []
            for col in range(3, 83, 2):
                value = sheet.cell(i, col).value
                data1.append(value)
            for col1 in range(4, 83, 2):
                value1 = sheet.cell(i + 1, col1).value
                data2.append(value1)
            for q in range(len(data1)):
                if data1[q] is None:
                    text = " "
                else:
                    text = f"{data1[q]}({data2[q]})"
                data.append(text)
        main_data = []
        for w in range(0, len(data), 8):
            txt = f"1.{data[w]}" \
                  f"\n2.{data[w + 1]}" \
                  f"\n3.{data[w + 2]}" \
                  f"\n4.{data[w + 3]}" \
                  f"\n5.{data[w + 4]}" \
                  f"\n6.{data[w + 5]}" \
                  f"\n7.{data[w + 6]}" \
                  f"\n8.{data[w + 7]}"
            main_data.append(txt)
        print(main_data)

        ids = []
        for i in range(1, tchr_count + 1):
            for day in range(1, 6):
                a = i + day / 10
                ids.append(a)

        for s in range(0, len(main_data)):
            self.cursor.execute(f"INSERT INTO uchitel_rasp VALUES (:id_day, :rasp);",
                                {'id_day': ids[s], 'rasp': main_data[s]})

        self.connect.commit()

    def kab(self, kab_count: int, min_row: int, max_row: int):
        sheet = self.file_to_read['кабинеты']
        data = []
        for i in range(min_row, max_row + 1):
            data1 = []
            for col in range(3, 43):
                value = sheet.cell(i, col).value
                data1.append(value)
            for q in range(len(data1)):
                if data1[q] is None:
                    text = " "
                else:
                    text = f"{data1[q]}"
                data.append(text)
        main_data = []
        for w in range(0, len(data), 8):
            txt = f"1.{data[w]}" \
                  f"\n2.{data[w + 1]}" \
                  f"\n3.{data[w + 2]}" \
                  f"\n4.{data[w + 3]}" \
                  f"\n5.{data[w + 4]}" \
                  f"\n6.{data[w + 5]}" \
                  f"\n7.{data[w + 6]}" \
                  f"\n8.{data[w + 7]}"
            main_data.append(txt)

        ids = []
        for i in range(1, kab_count + 1):
            for day in range(1, 6):
                a = i + day / 10
                ids.append(a)

        for s in range(0, len(main_data)):
            self.cursor.execute(f"INSERT INTO uchitel_kab VALUES (:id_day, :rasp);",
                                {'id_day': ids[s], 'rasp': main_data[s]})

        self.connect.commit()
